import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

st.title("Excel Cleaner and Pivot App")

# Upload Excel file
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    # Load all sheets
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names
    st.subheader("Sheets in uploaded file:")
    st.write(sheet_names)

    # Temporary buffer for output
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for i, sheet in enumerate(sheet_names):
            df = pd.read_excel(xls, sheet_name=sheet)
            st.subheader(f"Processing Sheet: {sheet}")
            st.dataframe(df)

            # Skip pivoting and formatting for the first sheet
            if i == 0:
                st.info(f"Skipping pivot for the first sheet: {sheet}")
                df.to_excel(writer, sheet_name=sheet, index=False)
                continue

            # Ensure there are at least 4 columns
            if len(df.columns) >= 4:
                # Original columns A–D
                original_part = df.iloc[:, :4]

                # Columns B–D for pivoting
                pivot_part = df.iloc[:, 1:4]

                # Combine: Original columns + pivoted columns starting at column E
                combined_df = pd.concat([original_part, pivot_part], axis=1)

                # Format columns B and E: remove '+', '.0', and NaN
                for col_index in [1, 4]:  # B and E are index 1 and 4 (0-based)
                    if col_index < len(combined_df.columns):
                        combined_df.iloc[:, col_index] = (
                            combined_df.iloc[:, col_index]
                            .astype(str)
                            .str.replace("+", "", regex=False)
                            .str.replace(".0", "", regex=False)
                            .str.strip()
                        )
                        combined_df.iloc[:, col_index] = combined_df.iloc[
                            :, col_index
                        ].replace(["nan", "NaN", "None"], "")
                        combined_df.iloc[:, col_index] = combined_df.iloc[
                            :, col_index
                        ].fillna("")
            else:
                combined_df = df  # fallback if not enough columns

            # Write modified sheet
            combined_df.to_excel(writer, sheet_name=sheet, index=False)

    # Apply formatting with openpyxl
    output.seek(0)
    workbook = load_workbook(output)

    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)

    for i, sheet_name in enumerate(sheet_names):
        ws = workbook[sheet_name]

        # Apply formatting only on sheets after the first
        if i != 0:
            # Highlight E1, F1, G1 yellow and bold
            for col in ["E", "F", "G"]:
                if ws[f"{col}1"].value is not None:
                    ws[f"{col}1"].fill = yellow_fill
                    ws[f"{col}1"].font = bold_font

        # Bold all headers (first row)
        for cell in ws[1]:
            cell.font = bold_font

        # Format all cells for readability
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
                if not cell.font.bold:  # keep headers bold, rest normal
                    cell.font = font

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    # Save formatted workbook
    final_output = BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    # Provide download button
    st.download_button(
        label="Download Formatted Cleaned Excel",
        data=final_output.getvalue(),
        file_name="formatted_cleaned_file.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
